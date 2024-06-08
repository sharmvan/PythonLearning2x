# # different types of ways to get the values in Excel sheet
# from openpyxl import Workbook, load_workbook
#
# lw = load_workbook(filename="demo.xlsx")
# sh = lw.active
# print(sh['A4'].value)
#
# from openpyxl import Workbook, load_workbook
#
# lw = load_workbook(filename="demo.xlsx")
# sh = lw['Sheet1']
# print(sh['A4'].value)
#
# from openpyxl import Workbook, load_workbook
#
# lw = load_workbook(filename="demo.xlsx")
# sh = lw['Sheet1']
# print(lw['Sheet1']['A4'].value)
#
# from openpyxl import Workbook, load_workbook
#
# lw = load_workbook(filename="demo.xlsx")
# sh = lw['Sheet1']
# print(sh.cell(4, 1).value)
#
# from openpyxl import Workbook, load_workbook
#
# lw = load_workbook(filename="demo.xlsx")
# sh = lw['Sheet1']
# print(sh.cell(row=4, column=1).value)  # More readable way to get the values in Excel sheet
#
# # Let's say we have a requirement to read the entire file and get all the values.
# # we need to know how many rows and columns are there in the excelsheet.
from openpyxl import Workbook, load_workbook

lw = load_workbook(filename="demo.xlsx")
sh = lw['Sheet1']
row_ct = sh.max_row
col_ct = sh.max_column
# print(row_ct)
# print(col_ct)
for i in range(1, row_ct + 1):
    for j in range(1, col_ct + 1):
        print(sh.cell(row=i, column=j).value, end=' ')
    print('\n')
