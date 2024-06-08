from openpyxl import Workbook

wb = Workbook()
ws = wb.active
# ws['F10'] = 'The Testing Academy'

testdata = [['Name', 'City'], ['Vandana', 'London'], ['Komal', 'India'], ['Anu', 'Paris']]
for data in testdata:
    ws.append(data)

# for i in range(1, 6):
#     for j in range(1, 6):
#         ws.cell(row=i, column=j).value = i + j
wb.save("demo.xlsx")
