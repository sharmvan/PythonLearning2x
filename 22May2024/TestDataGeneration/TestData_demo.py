# from faker import Faker
#
# fake_data = Faker()
#
# print(fake_data.name())
# print(fake_data.address())
# print(fake_data.email())
#
# print("----------------------------------------")
#
# from faker import Faker
# from openpyxl import Workbook
#
# wb = Workbook()
# sh = wb.active
# fake_data = Faker()
#
# for i in range(1, 11):
#     sh.cell(row=i, column=1).value = fake_data.name()
# wb.save("testData.xlsx")
#
# print("----------------------------------------")
#
# from faker import Faker
# from openpyxl import Workbook
#
# wb = Workbook()
# sh = wb.active
# fake_data = Faker()
#
# for i in range(1, 11):
#     for j in range(1, 4):
#         sh.cell(row=i, column=1).value = fake_data.name()
#         sh.cell(row=i, column=2).value = fake_data.city()
#         sh.cell(row=i, column=3).value = fake_data.email()
# wb.save("testData.xlsx")
#
# print("----------------------------------------")

# from faker import Faker
# from openpyxl import Workbook
#
# wb = Workbook()
# sh = wb.active
# fake_data = Faker('hi_IN') # Locale in Hindi
#
# for i in range(1, 11):
#     for j in range(1, 4):
#         sh.cell(row=i, column=1).value = fake_data.name()
#         sh.cell(row=i, column=2).value = fake_data.city()
#         sh.cell(row=i, column=3).value = fake_data.email()
# wb.save("testData.xlsx")
#
# print("----------------------------------------")
#
from faker import Faker
from openpyxl import Workbook

wb = Workbook()
sh = wb.active
fake_data = Faker(['hi_IN', 'en_US'])  # Locale (mix of Hindi & English)

for i in range(1, 11):
    for j in range(1, 4):
        sh.cell(row=i, column=1).value = fake_data.name()
        sh.cell(row=i, column=2).value = fake_data.city()
        sh.cell(row=i, column=3).value = fake_data.email()
wb.save("testData.xlsx")

print("----------------------------------------")
